#-------------------------------------------------------------------------------
# Name:			Mutant Codon to Amino Acid Organizer
# Purpose:		Checks nucleotide sequence for mutations and stores single amino acid substitions found in an excel spreadsheet
#
# Author:		Sarah Lach
#
# Created:		18/06/2019
# Copyright:	(c) Sarah Lach 2019
# Licence:		<your licence>
#-------------------------------------------------------------------------------

#This program reads in an excel spreadsheet provided by Genewiz ("plate1_abundance" or similarly named) to extract the data
#about the frequency of reads for mutated reads.
#It only retains data for reads with single codon change.
#The data is written to a new excel spreadsheet which is saved to the same location as the input file.
#The program must be run through a python virtual environment to respond to prompts.
#Stop mutants are marked as '*', silent mutations are tracked, marked by the same amino acid

import openpyxl
from openpyxl import Workbook
import os

#before beginning, you should filter your "plate1_abundance" spreadsheet to only have your desired number of changes and a read cutoff.
#you only need to save the first 10 columns, remove the top row with the column labels.
#going in, you will need some data on hand and available:
#the filepath to your filtered saved excel spreadsheet, the name of the spreadsheet, the number of reads before filtering,
#and how many bases from the start of the gene your reference begins, so the program can give the right frame & properly number residues.
#(if your reference sequence starts at the start of your gene, enter 0. If your reference sequence starts before the start of your gene, enter the number of nucleotides
#before the start of your gene as a negative number. Otherwise you need to enter the number of nucleotides in the gene before the start of your reference sequence.)



#Loading in the excel spreadsheet
print("give filepath to input workbook")
filepath = input()
print("filepath =", filepath)
os.chdir(filepath)
print("give name of workbook including .xlsx")
rawName = input()
print("workbook =", rawName)
raw = openpyxl.load_workbook(rawName)
print(raw.get_sheet_names())
print("from list of names, type which sheet you want to work with")
sheetName = input()
print("sheet =", sheetName)
rawSheet=raw.get_sheet_by_name(sheetName)


#Creating the output spreadsheet
out = Workbook()
outSheet = out.active
print("name output workbook")
title = input()
print("new workbook =", title)
outSheet.title = title
filename = title+'.xlsx'
out.save(filename)

#Adding column headers to the output spreadsheet
outSheet['A1'] = "change"
outSheet['B1'] = "original amino acid"
outSheet['C1'] = "location"
outSheet['D1'] = "new amino acid"
outSheet['E1'] = "reads (sum)"


#Entering the initial information needed for the program
print("input total number of reads (before filtering)")
readTot = int(input())
print("total reads =", readTot)
print("input offset number for ref")
seqOff = int(input())
print("offset =", seqOff)
AAOff = seqOff/3

outSheet['F1'] = "reads/"+str(readTot)


#Get the reference sequence from the spreadsheet and find its length
refSCell = rawSheet.cell(row=2, column=10)
refS = refSCell.value
leng = len(refS)


#method that takes in the reference sequence, the sequence for the read, and the length of the reference sequence to find up to 3 nucleotide variants, by codon
def codonIter(rSeq, tSeq, refLeng):
	#if the reference does not start at the beginning of a codon, adjusts the number to look only at codons as a whole
	if seqOff%3!=0:
		frontFrame = 3-seqOff%3
		rSeq = rSeq[frontFrame:]
		refLeng-=frontFrame
	if refLeng%3!=0:
		backFrame = refLeng%3
		rSeq = rSeq[0:refLeng-backFrame]
		refLeng-=backFrame

	#variables for the nucleotide number for the mutant codon
	locS1 = -1
	locS2 = -1
	locS3 = -1

	#variables for the mutant codon
	m1 = ''
	m2 = ''
	m3 = ''

	#variables for the reference codon at the same location as the mutant
	rb1 = ''
	rb2 = ''
	rb3 = ''

	#loop to iterate through the mutant and reference sequence, comparing the two, and when differences are found, saving the location, mutant codon, and reference codon
	i=0
	while i<refLeng:
		if rSeq[i:i+3]==tSeq[i:i+3]:
			i+=3
			continue
		else:
			if locS1==-1:
				locS1 = i+seqOff
				m1 = tSeq[i:i+3]
				rb1 = rSeq[i:i+3]
			elif locS2==-1:
				locS2 = i+seqOff
				m2 = tSeq[i:i+3]
				rb2 = rSeq[i:i+3]
			elif locS3==-1:
				locS3 = i+seqOff
				m3 = tSeq[i:i+3]
				rb3 = rSeq[i:i+3]
			i+=3
	return rb1,locS1,m1, rb2,locS2,m2, rb3,locS3,m3

codonDict = {'ttt': 'F','ttc': 'F','tta': 'L','ttg': 'L','ctt': 'L','ctc': 'L','cta': 'L','ctg': 'L','att': 'I','atc': 'I','ata': 'I','atg': 'M','gtt': 'V','gtc': 'V','gta': 'V','gtg': 'V','tct': 'S','tcc': 'S','tca': 'S','tcg': 'S','cct': 'P','ccc': 'P','cca': 'P','ccg': 'P','act': 'T','acc': 'T','aca':'T','acg':'T','gct':'A','gcc':'A','gca': 'A','gcg': 'A','tat': 'Y','tac': 'Y','taa': '*','tag': '*','cat': 'H','cac': 'H','caa': 'Q','cag': 'Q','aat': 'N','aac': 'N','aaa': 'K','aag': 'K','gat': 'D','gac': 'D','gaa':'E','gag':'E','tgt':'C','tgc':'C','tga': '*','tgg': 'W','cgt': 'R','cgc': 'R','cga': 'R','cgg': 'R','agt': 'S','agc': 'S','aga': 'R','agg':'R','ggt':'G','ggc':'G','gga':'G','ggg':'G'}
doubleNum=1

#searches to see if a specific amino acid substitution has already been added to the output excel sheet
def changeSearch(search):
	for row_index in range(1, outSheet.max_row+1):
		if outSheet.cell(row = row_index,column=1).value == search:
			return row_index
	return row_index+1

#loop to iterate through the rows of the spreadsheet.
for r in range(1, rawSheet.max_row+1):

	#get the number of reads this sequence was found in
	readCell = rawSheet.cell(row=r, column=2)
	readNum = int(readCell.value)

	#calculate the fraction of reads this sequence was found in
	percentNum = float(readNum/readTot)

	#get the sequence for this row
	targSeqCell = rawSheet.cell(row=r, column=9)
	targSeq = targSeqCell.value

	#get the reference sequence
	refSeqCell = rawSheet.cell(row=r, column=10)
	refSeq = refSeqCell.value

	#variables for the residue number for the nucleotide variant
	locA1 = -1
	locA2 = -1
	locA3 = -1

	#variables for the translated amino acid from nucleotide variant
	a1 = ''
	a2 = ''
	a3 = ''

	#variables for the translated amino acid from the reference sequence at the same location as the nucleotide variant
	ra1 = ''
	ra2 = ''
	ra3 = ''

	#getting the mutant codon, nucleotide number, and reference codon (for up to 3 codons) from codonIter.
	unAltC1,sL1, unAltM1,unAltC2,sL2, unAltM2,unAltC3,sL3, unAltM3 = codonIter(refSeq,targSeq,leng)

	#translating the nucleotide variant information into amino acid substitution information
	if(sL1!=-1):
		locA1 = int(sL1/3)+1
		ra1 = codonDict[unAltC1.lower()]
		a1 = codonDict[unAltM1.lower()]
	if(sL2!=-1):
		locA2 = int(sL2/3)+1
		ra2 = codonDict[unAltC2.lower()]
		a2 = codonDict[unAltM2.lower()]
	if(sL3!=-1):
		locA3 = int(sL3/3)+1
		ra3 = codonDict[unAltC3.lower()]
		a3 = codonDict[unAltM3.lower()]

	#recording the amino acid substitution found when only 1 amino acid substitution was detected.
	#If the amino acid substitution is found in multiple sequences, combines the data
	if(locA1!=-1):
		if(locA2==-1):
			change = ra1+str(locA1)+a1
			curRow = changeSearch(change)
			writeChange = outSheet.cell(row=curRow, column=1)
			writeChange.value = change

			writeStartAA = outSheet.cell(row=curRow, column =2)
			writeStartAA.value = ra1
			writeLoc = outSheet.cell(row=curRow, column=3)
			writeLoc.value = locA1
			writeFinalAA = outSheet.cell(row=curRow, column=4)
			writeFinalAA.value = a1

			writeReads = outSheet.cell(row=curRow, column=5)
			if writeReads.value == None:
				writeReads.value = readNum
			else:
				writeReads.value += readNum
			writePercent = outSheet.cell(row=curRow, column=6)
			if writePercent.value == None:
				writePercent.value = percentNum
			else:
				writePercent.value += percentNum
			out.save(filename)


